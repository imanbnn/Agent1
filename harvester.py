import asyncio
import os
import re
import json
import sys
import subprocess
import datetime
import sqlite3
from logger import log_event
from scroll import force_scroll_down
from actions import teach_and_click, gemini_double_check_totals, get_supervisor_status, dismiss_all_popups

try:
    from catalog_data import GFS_CATALOG_TREE
except ImportError:
    GFS_CATALOG_TREE = {}

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("📦 [SYSTEM] Installing 'openpyxl' for Excel Master Sheet generation...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
    import openpyxl
    from openpyxl.styles import Font, Alignment

# 🛡️ ONE-TIME ERASE: Bumping to v4 forces a brand new, clean database with the split columns
DB_FILE = "gfs_master_v4.db"
EXCEL_FILE = "GFS_Master_Catalog.xlsx"
product_db = {}
target_total = 9999 
RAW_INFO_BUFFER = []
RAW_PRICE_BUFFER = []
NO_IMAGE_SVG = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNjAiIGhlaWdodD0iMTYwIj48cmVjdCB3aWR0aD0iMTAwJSIgaGVpZ2h0PSIxMDAlIiBmaWxsPSIjZWVlIi8+PHRleHQgeD0iNTAlIiB5PSI1MCUiIGZvbnQtZmFtaWx5PSJzYW5zLXNlcmlmIiBmb250LXNpemU9IjE0IiBmb250LXdlaWdodD0iYm9sZCIgZmlsbD0iIzk5OSIgdGV4dC1hbmNob3I9Im1pZGRsZSIgZHk9Ii4zZW0iPkltYWdlIE5vdCBGb3VuZDwvdGV4dD48L3N2Zz4="

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS products (
            code TEXT PRIMARY KEY,
            name TEXT,
            brand TEXT,
            cat_l1 TEXT,
            cat_l2 TEXT,
            cat_l3 TEXT,
            measure_1 TEXT,
            measure_2 TEXT,
            price_1_unit TEXT,
            price_1 TEXT,
            price_2_unit TEXT,
            price_2 TEXT,
            stock TEXT,
            last_ordered TEXT,
            image_file TEXT,
            img_url TEXT,
            changes TEXT,
            last_updated TEXT,
            last_seen TEXT
        )
    ''')
    conn.commit()
    conn.close()

def process_changes():
    init_db()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    c.execute("SELECT code, price_1, price_2, img_url, changes, last_updated FROM products")
    existing_products = {row[0]: row[1:] for row in c.fetchall()}
    
    inserts = []
    updates = []
    
    for code, details in product_db.items():
        if code in existing_products:
            old_p1, old_p2, old_img, old_changes, old_last_updated = existing_products[code]
            has_changed = False
            
            if old_p1 != details["price_1"] and details["price_1"]:
                change_msg = f"Primary Price Changed to {details['price_1']}"
                has_changed = True
            elif old_p2 != details["price_2"] and details["price_2"]:
                change_msg = f"Secondary Price Changed to {details['price_2']}"
                has_changed = True
            elif old_img != details["img_url"] and details["img_url"]:
                change_msg = f"New Image"
                has_changed = True
            else:
                change_msg = old_changes or ""
                
            last_updated_val = current_time if has_changed else (old_last_updated or current_time)
            details["changes"] = change_msg
            
            updates.append((
                details["name"], details["brand"], details["cat_l1"], details["cat_l2"], details["cat_l3"],
                details["measure_1"], details["measure_2"], details["price_1_unit"], 
                details["price_1"], details["price_2_unit"], details["price_2"], 
                details["stock"], details["last_ordered"], details["image_file"], 
                details["img_url"], change_msg, last_updated_val, current_time, code
            ))
        else:
            change_msg = f"New Product"
            details["changes"] = change_msg
            last_updated_val = current_time
            
            inserts.append((
                code, details["name"], details["brand"], details["cat_l1"], details["cat_l2"], details["cat_l3"],
                details["measure_1"], details["measure_2"], details["price_1_unit"], 
                details["price_1"], details["price_2_unit"], details["price_2"], 
                details["stock"], details["last_ordered"], details["image_file"], 
                details["img_url"], change_msg, last_updated_val, current_time
            ))
            
    if updates:
        c.executemany("""
            UPDATE products 
            SET name=?, brand=?, cat_l1=?, cat_l2=?, cat_l3=?, measure_1=?, measure_2=?, price_1_unit=?, price_1=?, price_2_unit=?, price_2=?, stock=?, last_ordered=?, image_file=?, img_url=?, changes=?, last_updated=?, last_seen=?
            WHERE code=?
        """, updates)
        
    if inserts:
        c.executemany("""
            INSERT INTO products 
            (code, name, brand, cat_l1, cat_l2, cat_l3, measure_1, measure_2, price_1_unit, price_1, price_2_unit, price_2, stock, last_ordered, image_file, img_url, changes, last_updated, last_seen)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, inserts)
        
    conn.commit()
    conn.close()

def generate_excel_master():
    print(f"📊 Generating perfectly formatted Excel Master Catalog: {EXCEL_FILE}...")
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT code, name, cat_l1, cat_l2, cat_l3, brand, measure_1, measure_2, price_1_unit, price_1, price_2_unit, price_2, stock, last_ordered, changes, last_updated, last_seen FROM products ORDER BY name ASC")
        rows = c.fetchall()
        conn.close()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Catalog"
        
        headers = ["Code", "Product Name", "Department (L1)", "Category (L2)", "Sub-Category (L3)", "Brand", "Measure 1 (Weight/Vol)", "Measure 2 (Base UOM)", "Primary Unit", "Primary Price", "Secondary Unit", "Secondary Price", "Stock Status", "Last Ordered", "Recent Changes", "Last Updated", "Last Seen"]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:Q{len(rows) + 1}"
        
        for row_data in rows:
            ws.append(row_data)
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col[:100]: 
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min(max_length + 2, 50) 
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(EXCEL_FILE)
        print("✅ Excel Master Catalog successfully updated!")
    except Exception as e:
        print(f"⚠️ Failed to generate Excel file: {e}")

def update_status(state, message, current=0, total=0):
    try:
        with open("status.json", "w", encoding="utf-8") as f:
            json.dump({"state": state, "message": message, "current": current, "total": total}, f)
    except: pass

def generate_html_dashboard():
    print("🌐 Generating Lightning-Fast HTML Visual Dashboard...")
    sup_status = get_supervisor_status()
    raw_data = []
    
    try:
        if os.path.exists(DB_FILE):
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT name, code, cat_l1, cat_l2, cat_l3, brand, measure_1, measure_2, price_1_unit, price_1, price_2_unit, price_2, stock, last_ordered, image_file, changes, last_updated, last_seen FROM products ORDER BY last_seen DESC")
            rows = c.fetchall()
            
            for row in rows:
                name, code, l1, l2, l3, brand, m1, m2, p1u, p1, p2u, p2, stock, last_ord, img, changes, last_updated, last_seen = row
                img_path = img if img not in ["No Image", "Failed to Download", ""] else NO_IMAGE_SVG
                
                raw_data.append({
                    "code": code,
                    "name": name,
                    "l1": l1,
                    "l2": l2,
                    "l3": l3,
                    "brand": brand,
                    "m1": m1,
                    "m2": m2,
                    "p1u": p1u,
                    "p1": p1,
                    "p2u": p2u,
                    "p2": p2,
                    "stock": stock,
                    "last_ord": last_ord,
                    "img": img_path,
                    "changes": changes,
                    "last_updated": last_updated,
                    "last_seen": last_seen
                })
            conn.close()
    except Exception as e:
        print(f"⚠️ Failed to read DB for dashboard: {e}")
        
    tree_json = json.dumps(GFS_CATALOG_TREE).replace("'", "\\'").replace("\\\\", "\\\\\\\\")
        
    html_content = f"""<!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>GFS Harvest Dashboard</title>
        <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
        <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
        <script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f4f7f6; padding: 140px 20px 20px; color: #333; margin: 0; }}
            .control-panel {{ position: fixed; top: 0; left: 0; right: 0; background: white; padding: 15px 30px; box-shadow: 0 4px 10px rgba(0,0,0,0.1); z-index: 1000; display: flex; flex-direction: column; gap: 15px; }}
            .panel-top {{ display: flex; align-items: center; gap: 20px; justify-content: space-between; width: 100%; }}
            .panel-left {{ display: flex; align-items: center; gap: 15px; flex-grow: 1; }}
            .btn-search, .btn-explore, .btn-view {{ color: white; border: none; padding: 10px 20px; font-size: 14px; font-weight: bold; border-radius: 6px; cursor: pointer; transition: all 0.2s; box-shadow: 0 4px 6px rgba(0,0,0,0.15); }}
            .btn-search {{ background: #27ae60; }} .btn-search:hover {{ background: #219150; transform: translateY(-1px); }}
            .btn-explore {{ background: #8e44ad; }} .btn-explore:hover {{ background: #732d91; transform: translateY(-1px); }}
            .btn-view {{ background: #ecf0f1; color: #2c3e50; box-shadow: none; border: 1px solid #bdc3c7; }} 
            .btn-view:hover {{ background: #bdc3c7; }}
            .btn-search:disabled, .btn-explore:disabled {{ background: #95a5a6; cursor: not-allowed; box-shadow: none; transform: none; }}
            .view-toggle-container {{ display: flex; gap: 10px; border-left: 2px solid #eee; padding-left: 20px; margin-left: 10px; }}
            .active-view {{ background: #34495e !important; color: white !important; border-color: #34495e !important; box-shadow: 0 4px 6px rgba(52, 73, 94, 0.3) !important; }}
            .status-container {{ flex-grow: 1; max-width: 500px; margin-left: 20px; }}
            .progress-bg {{ background: #ecf0f1; height: 12px; border-radius: 6px; overflow: hidden; width: 100%; box-shadow: inset 0 1px 3px rgba(0,0,0,0.1); }}
            .supervisor-badge {{ background: #2c3e50; color: white; padding: 10px 20px; border-radius: 8px; font-weight: bold; font-size: 0.9em; white-space: nowrap; }}
            h1 {{ text-align: center; color: #2c3e50; margin-top: 0; }}
            
            .grid {{ display: none; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 20px; padding-bottom: 20px; margin-top: 15px; }}
            .card {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); display: flex; flex-direction: column; transition: transform 0.2s; position: relative; border-top: 4px solid transparent; }}
            .card:hover {{ transform: translateY(-5px); box-shadow: 0 8px 15px rgba(0,0,0,0.15); }}
            .card img {{ max-width: 100%; height: 160px; object-fit: contain; align-self: center; margin-bottom: 15px; border-radius: 4px; }}
            .card h3 {{ margin: 0 0 10px 0; font-size: 1.1em; color: #2c3e50; line-height: 1.3; }}
            .details {{ color: #7f8c8d; font-size: 0.9em; margin-bottom: 15px; flex-grow: 1; line-height: 1.5; }}
            .price {{ color: #27ae60; font-weight: bold; font-size: 1.3em; margin-bottom: 5px; }}
            .stock {{ font-weight: bold; font-size: 0.9em; margin-bottom: 5px; }}
            .in-stock {{ color: #27ae60; }} 
            .out-of-stock {{ color: #e74c3c; }} 
            .contact-sales {{ color: #f39c12; }}
            .card.out-of-stock {{ border-top-color: #e74c3c; }}
            .ordered {{ color: #34495e; font-size: 0.85em; margin-top: 10px; padding-top: 10px; border-top: 1px solid #eee; }}
            .badge {{ background: #e74c3c; color: white; padding: 4px 10px; border-radius: 12px; font-size: 0.8em; font-weight: bold; margin-bottom: 12px; display: inline-block; align-self: flex-start; box-shadow: 0 2px 4px rgba(231, 76, 60, 0.3); }}
            .cat-badge {{ background: #ecf0f1; color: #2c3e50; padding: 2px 6px; border-radius: 4px; font-size: 0.85em; border: 1px solid #bdc3c7; display: inline-block; margin: 2px 0; }}
            
            .dataTables_wrapper {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin-top: 20px; }}
            table.dataTable thead th {{ background-color: #34495e; color: white; padding: 12px 10px; border-bottom: none; }}
            table.dataTable tbody tr:hover {{ background-color: #f8f9fa; }}
            table.dataTable tbody td {{ vertical-align: middle; }}
            .dataTables_wrapper .dataTables_filter input {{ border: 1px solid #bdc3c7; border-radius: 4px; padding: 8px 12px; margin-left: 10px; width: 250px; }}
            
            /* CSS GHOSTING */
            body.grid-mode table.dataTable {{ position: absolute !important; visibility: hidden !important; height: 0 !important; overflow: hidden !important; border: none !important; }}
            body.grid-mode .grid {{ display: grid !important; }}
            body.grid-mode .dataTables_wrapper {{ background: transparent; box-shadow: none; padding: 0; }}

            /* Dynamic Tree UI CSS */
            .modal-backdrop {{ display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: rgba(0,0,0,0.6); z-index: 1999; backdrop-filter: blur(3px); }}
            .category-modal {{ display: none; position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%); background: white; padding: 30px; border-radius: 12px; box-shadow: 0 15px 35px rgba(0,0,0,0.2); z-index: 2000; width: 600px; max-width: 90%; }}
            .tree-container {{ max-height: 400px; overflow-y: auto; margin: 15px 0; border: 1px solid #eee; padding: 15px; border-radius: 6px; background: #fdfdfd; }}
            ul.tree-list {{ list-style-type: none; padding-left: 20px; margin-top: 5px; }}
            ul.tree-root {{ padding-left: 0; }}
            .tree-list li {{ margin-bottom: 8px; }}
            .tree-list label {{ cursor: pointer; display: flex; align-items: center; gap: 8px; font-size: 15px; color: #2c3e50; }}
            .tree-list input[type="checkbox"] {{ transform: scale(1.2); cursor: pointer; }}
            .dept-label {{ font-size: 18px !important; font-weight: bold; border-bottom: 1px solid #ddd; padding-bottom: 5px; margin-top: 15px; display: block; }}
            .sub-label {{ font-size: 16px !important; font-weight: 600; color: #34495e; }}
            .global-select-container {{ background: #ecf0f1; padding: 15px; border-radius: 6px; margin-bottom: 15px; border: 1px solid #bdc3c7; }}
        </style>
    </head>
    <body>
        <div class="control-panel">
            <div class="panel-top">
                <div class="panel-left">
                    <button id="searchBtn" class="btn-search" onclick="openModal()">📋 Select URL Targets to Scrape</button>
                    <button id="exploreBtn" class="btn-explore" onclick="startBot('explore')">🧭 Brute-Force Entire Catalog</button>
                    <div class="view-toggle-container">
                        <button id="btn-grid" class="btn-view active-view" onclick="toggleView('grid')">📱 Grid View</button>
                        <button id="btn-table" class="btn-view" onclick="toggleView('table')">🗄️ ERP Database View</button>
                    </div>
                    <div class="status-container">
                        <p id="status-text">Status: Ready</p>
                        <div class="progress-bg"><div id="progress-bar"></div></div>
                    </div>
                </div>
                <div class="supervisor-badge">🧠 Supervisor: {sup_status}</div>
            </div>
        </div>

        <div id="modalBackdrop" class="modal-backdrop" onclick="closeModal()"></div>
        <div id="categoryModal" class="category-modal">
            <h2 style="margin-top: 0; color: #2c3e50;">🎯 Select Deep Catalog Targets</h2>
            <p style="color: #7f8c8d; font-size: 0.9em; margin-top: -10px;">Select the specific sub-categories you want the bot to navigate to.</p>
            
            <div class="global-select-container">
                <label style="font-size: 16px; font-weight: bold; cursor: pointer; display: flex; align-items: center; gap: 10px;">
                    <input type="checkbox" id="selectAllGlobal" style="transform: scale(1.4);"> 
                    ✅ SELECT ALL AVAILABLE CATEGORIES
                </label>
            </div>

            <div class="tree-container" id="treeRenderer">
                </div>
            
            <div style="display:flex; justify-content:flex-end; gap:10px;">
                <button onclick="closeModal()" class="btn-view">Cancel</button>
                <button onclick="startSelectedHarvest()" class="btn-search">🚀 Start Custom Harvest</button>
            </div>
        </div>

        <h1>📦 Master Local Database Dashboard</h1>
        
        <table id="erpTable" class="display" style="width:100%">
            <thead>
                <tr>
                    <th>Image</th>
                    <th>Code</th>
                    <th>Name</th>
                    <th>Dept (L1)</th>
                    <th>Category (L2)</th>
                    <th>Sub-Cat (L3)</th>
                    <th>Brand</th>
                    <th>Price 1</th>
                    <th>Price 2</th>
                    <th>Stock Status</th>
                    <th>Last Ordered</th>
                    <th>Recent Changes</th>
                    <th>Last Seen</th>
                </tr>
            </thead>
            <tbody>
            </tbody>
        </table>
        
        <script>
            const rawData = {json.dumps(raw_data)};
            const noImageSvg = '{NO_IMAGE_SVG}';
            const catalogTree = JSON.parse('{tree_json}');
            
            $(document).ready(function() {{
                const table = $('#erpTable').DataTable({{
                    data: rawData,
                    pageLength: 25,
                    lengthMenu: [10, 25, 50, 100, 250],
                    deferRender: true,
                    searchDelay: 400, 
                    order: [[12, 'desc']], 
                    columns: [
                        {{ data: 'img', render: function(data) {{ return `<img src='${{data}}' loading="lazy" style='width: 45px; height: 45px; object-fit: contain; border-radius: 4px;' onerror="this.src='${{noImageSvg}}'">`; }}, orderable: false }},
                        {{ data: 'code', render: function(data) {{ return `<b>#${{data}}</b>`; }} }},
                        {{ data: 'name' }},
                        {{ data: 'l1', render: function(data) {{ return `<span style="color:#8e44ad; font-weight:bold;">${{data || '-'}}</span>`; }} }},
                        {{ data: 'l2', render: function(data) {{ return `<span style="color:#2980b9; font-weight:bold;">${{data || '-'}}</span>`; }} }},
                        {{ data: 'l3' }},
                        {{ data: 'brand' }},
                        {{ data: null, render: function(data, type, row) {{ return row.p1 ? `<span style='color: #27ae60; font-weight: bold;'>${{row.p1u || 'Unit'}}: ${{row.p1}}</span>` : 'Price Not Found'; }} }},
                        {{ data: null, render: function(data, type, row) {{ return row.p2 ? `<span style='color: #2980b9; font-weight: bold;'>${{row.p2u || 'Unit'}}: ${{row.p2}}</span>` : ''; }} }},
                        {{ data: 'stock', render: function(data) {{
                            let sc = "in-stock";
                            let rawStock = data ? data.toUpperCase() : "";
                            if(rawStock.includes("OUT OF STOCK")) sc = "out-of-stock";
                            else if(rawStock.includes("CONTACT")) sc = "contact-sales";
                            return `<span class='stock ${{sc}}'>${{data || 'In Stock'}}</span>`;
                        }}}},
                        {{ data: 'last_ord' }},
                        {{ data: 'changes' }},
                        {{ data: 'last_seen' }}
                    ],
                    language: {{
                        search: "Global Filter (Searches Grid & Table instantly):",
                        lengthMenu: "Show _MENU_ products per page"
                    }}
                }});

                $('<div id="grid-container" class="grid"></div>').insertAfter('#erpTable');

                function renderGrid() {{
                    const grid = $('#grid-container');
                    let htmlArray = [];
                    
                    table.rows({{ page: 'current' }}).every(function() {{
                        let r = this.data();
                        
                        let rawStock = r.stock ? r.stock.toUpperCase() : "";
                        let sc = "in-stock";
                        if(rawStock.includes("OUT OF STOCK")) sc = "out-of-stock";
                        else if(rawStock.includes("CONTACT")) sc = "contact-sales";

                        let badge = r.changes ? `<div class="badge">🆕 ${{r.changes}}</div>` : "";
                        let price1 = r.p1 ? `${{r.p1u || 'Unit'}}: ${{r.p1}}` : "Price Not Found";
                        let price2 = r.p2 ? ` | ${{r.p2u || 'Unit'}}: ${{r.p2}}` : "";
                        
                        let detailsArr = [];
                        if (r.brand && r.brand !== 'Unknown') detailsArr.push(r.brand);
                        if (r.m1) detailsArr.push(r.m1);
                        if (r.m2) detailsArr.push(r.m2);
                        let details = detailsArr.join(' | ');
                        
                        let breadcrumbBadges = "";
                        if (r.l1) breadcrumbBadges += `<span class="cat-badge" style="background:#f3e5f5; border-color:#e1bee7;">${{r.l1}}</span> `;
                        if (r.l2) breadcrumbBadges += `<span class="cat-badge" style="background:#e3f2fd; border-color:#bbdefb;">${{r.l2}}</span> `;
                        if (r.l3) breadcrumbBadges += `<span class="cat-badge">${{r.l3}}</span>`;

                        htmlArray.push(`
                        <div class="card ${{sc}}">
                            ${{badge}}
                            <img src="${{r.img || noImageSvg}}" loading="lazy" alt="Product Image" onerror="this.src='${{noImageSvg}}'">
                            <h3>${{r.name || 'Unknown Product'}}</h3>
                            <div class="details">
                                <b>#${{r.code}}</b><br>
                                <div style="margin: 8px 0;">${{breadcrumbBadges}}</div>
                                ${{details}}
                            </div>
                            <div class="price">${{price1}}${{price2}}</div>
                            <div class="stock ${{sc}}">${{r.stock || 'In Stock'}}</div>
                            <div class="ordered">Last Ordered: ${{r.last_ord || 'N/A'}} <br><small>Updated: ${{r.last_updated}}</small></div>
                        </div>`);
                    }});
                    
                    grid.html(htmlArray.join(''));
                }}

                table.on('draw', renderGrid);
                renderGrid(); 

                window.toggleView = function(viewType) {{
                    if (viewType === 'grid') {{
                        $('body').addClass('grid-mode');
                        $('#btn-grid').addClass('active-view');
                        $('#btn-table').removeClass('active-view');
                    }} else {{
                        $('body').removeClass('grid-mode');
                        $('#btn-table').addClass('active-view');
                        $('#btn-grid').removeClass('active-view');
                        table.columns.adjust().draw(false); 
                    }}
                }};

                toggleView('grid');
                buildTreeUI();
            }});

            // 🌳 Dynamically generate the checkable tree UI
            function buildTreeUI() {{
                let html = '<ul class="tree-list tree-root">';
                
                for (let dept in catalogTree) {{
                    let safeDeptId = dept.replace(/[\\s&]+/g, '-');
                    html += `<li>
                        <label class="dept-label">
                            <input type="checkbox" class="dept-cb" data-target="${{safeDeptId}}"> 
                            ${{dept}}
                        </label>
                        <ul class="tree-list" id="${{safeDeptId}}">`;
                        
                    for (let sub in catalogTree[dept].nodes) {{
                        let node = catalogTree[dept].nodes[sub];
                        let safeSubId = sub.replace(/[\\s&]+/g, '-');
                        
                        html += `<li>
                            <label class="sub-label">
                                <input type="checkbox" class="sub-cb" data-target="${{safeSubId}}" value="${{node.url}}"> 
                                ${{sub}}
                            </label>
                            <ul class="tree-list" id="${{safeSubId}}">`;
                            
                        for (let leaf in node.leaves) {{
                            let leafUrl = node.leaves[leaf];
                            html += `<li>
                                <label>
                                    <input type="checkbox" class="leaf-cb" value="${{leafUrl}}"> 
                                    ${{leaf}}
                                </label>
                            </li>`;
                        }}
                        html += `</ul></li>`;
                    }}
                    html += `</ul></li>`;
                }}
                html += '</ul>';
                
                if(Object.keys(catalogTree).length === 0) {{
                    html = '<p style="color:red; font-weight:bold;">⚠️ The Catalog Brain is empty! Please verify your catalog_data.py file.</p>';
                }}
                
                $('#treeRenderer').html(html);

                // 🔄 Event Listeners for cascading checks
                $('#selectAllGlobal').on('change', function() {{
                    $('#treeRenderer input[type="checkbox"]').prop('checked', this.checked);
                }});

                $('.dept-cb').on('change', function() {{
                    let targetId = $(this).data('target');
                    $(`#${{targetId}} input[type="checkbox"]`).prop('checked', this.checked);
                }});

                $('.sub-cb').on('change', function() {{
                    let targetId = $(this).data('target');
                    $(`#${{targetId}} input[type="checkbox"]`).prop('checked', this.checked);
                }});
            }}

            function openModal() {{
                document.getElementById('modalBackdrop').style.display = 'block';
                document.getElementById('categoryModal').style.display = 'block';
            }}
            function closeModal() {{
                document.getElementById('modalBackdrop').style.display = 'none';
                document.getElementById('categoryModal').style.display = 'none';
            }}

            // 🏎️ SPEED OPTIMIZATION: Send the Parent URL if all children are checked
            function startSelectedHarvest() {{
                let selectedUrls = [];
                
                $('.sub-cb').each(function() {{
                    let subId = $(this).data('target');
                    let allLeaves = $(`#${{subId}} .leaf-cb`);
                    let checkedLeaves = $(`#${{subId}} .leaf-cb:checked`);
                    
                    if (allLeaves.length > 0) {{
                        if (allLeaves.length === checkedLeaves.length) {{
                            // ALL children checked! Send the single Level 2 parent URL to scrape instantly.
                            selectedUrls.push($(this).val());
                        }} else {{
                            // Only SOME children checked. Send the specific Level 3 leaf URLs.
                            checkedLeaves.each(function() {{
                                selectedUrls.push($(this).val());
                            }});
                        }}
                    }} else if ($(this).is(':checked')) {{
                        // No children exist, and it's checked. Send the URL.
                        selectedUrls.push($(this).val());
                    }}
                }});

                if (selectedUrls.length === 0) {{
                    alert("Please select at least one category to scrape!");
                    return;
                }}

                selectedUrls = [...new Set(selectedUrls)];

                document.getElementById('searchBtn').disabled = true;
                document.getElementById('exploreBtn').disabled = true;
                closeModal();
                document.getElementById('status-text').innerText = `Status: Initializing ${{selectedUrls.length}} target URLs...`;

                fetch('/api/start', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{ mode: 'custom_search', urls: selectedUrls }})
                }}).then(() => {{
                    console.log("Custom URL queue submitted to backend successfully.");
                }});
            }}

            function startBot(mode) {{
                document.getElementById('searchBtn').disabled = true;
                document.getElementById('exploreBtn').disabled = true;
                fetch('/api/start', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{ mode: mode, urls: [] }})
                }}).then(() => {{
                    document.getElementById('status-text').innerText = "Status: Initializing " + mode.toUpperCase() + " mode...";
                }});
            }}
            
            setInterval(() => {{
                fetch('status.json?t=' + Date.now())
                    .then(r => r.json())
                    .then(data => {{
                        document.getElementById('status-text').innerText = "Status: " + data.message;
                        let pct = data.total > 0 ? (data.current / data.total) * 100 : 0;
                        if(data.state === "complete") pct = 100;
                        document.getElementById('progress-bar').style.width = pct + '%';
                        if(data.state === 'running' || data.state === 'processing') {{
                            document.getElementById('searchBtn').disabled = true;
                            document.getElementById('exploreBtn').disabled = true;
                        }} else {{
                            document.getElementById('searchBtn').disabled = false;
                            document.getElementById('exploreBtn').disabled = false;
                        }}
                        if(data.state === 'complete' && !window.hasRefreshed) {{
                            window.hasRefreshed = true;
                            setTimeout(() => window.location.reload(), 2500);
                        }}
                        if(data.state === 'idle') window.hasRefreshed = false;
                    }}).catch(e => {{}});
            }}, 1500);
        </script>
    </body></html>
    """
    try:
        with open("dashboard.html", "w", encoding="utf-8") as f:
            f.write(html_content)
    except Exception as e:
        print(f"⚠️ Failed to generate HTML dashboard: {e}")

async def download_pending_images(page):
    os.makedirs("dashboard_images", exist_ok=True)
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "image/avif,image/webp,image/*,*/*;q=0.8"}
    pending = [k for k, v in product_db.items() if v.get("img_url") and v.get("image_file") in ["No Image", "Failed to Download"]]
    
    if not pending: return
    completed = 0
    sem = asyncio.Semaphore(15) 
    
    async def fetch_img(code, details):
        nonlocal completed
        img_file = f"dashboard_images/{code}.jpg"
        if not os.path.exists(img_file):
            try:
                async with sem:
                    url = details["img_url"]
                    if url.startswith("//"): url = "https:" + url
                    response = await page.request.get(url, headers=headers, timeout=10000)
                    if response.ok:
                        with open(img_file, 'wb') as f: f.write(await response.body())
                        product_db[code]["image_file"] = img_file
            except: product_db[code]["image_file"] = "Failed to Download"
        else: product_db[code]["image_file"] = img_file
        completed += 1
        if completed % 10 == 0: update_status("running", f"Downloading Images... {completed}/{len(pending)}", completed, len(pending))
        
    print(f"📸 Downloading {len(pending)} queued images in parallel...")
    await asyncio.gather(*[fetch_img(code, product_db[code]) for code in pending])

async def setup_wiretap(page):
    global RAW_INFO_BUFFER, RAW_PRICE_BUFFER
    RAW_INFO_BUFFER.clear()
    RAW_PRICE_BUFFER.clear()
    
    async def handle_response(response):
        global target_total, RAW_INFO_BUFFER, RAW_PRICE_BUFFER
        url = response.url.lower()
        if response.request.resource_type in ["fetch", "xhr"]:
            try:
                data = await response.json()
                if "totalResults" in data and ("materials/search" in url or "search" in url): 
                    new_total = int(data["totalResults"])
                    if target_total == 9999 or new_total > target_total:
                        target_total = new_total
                if "materialInfos" in data:
                    RAW_INFO_BUFFER.append(data)
                if "materialPrices" in data:
                    RAW_PRICE_BUFFER.append(data)
            except: pass
            
    page.on("response", handle_response)

def split_breadcrumb(breadcrumb_str):
    """Splits 'Categories > L1 > L2 > L3' into distinct columns."""
    if not breadcrumb_str or breadcrumb_str == "Unknown Category":
        return "", "", ""
        
    parts = [p.strip() for p in breadcrumb_str.split(">")]
    if parts[0].lower() == "categories":
        parts = parts[1:]
        
    l1 = parts[0] if len(parts) > 0 else ""
    l2 = parts[1] if len(parts) > 1 else ""
    l3 = parts[2] if len(parts) > 2 else ""
    
    return l1, l2, l3

def merge_buffers_to_db(page_breadcrumbs="Unknown Category"):
    global product_db, RAW_INFO_BUFFER, RAW_PRICE_BUFFER
    
    cat_l1, cat_l2, cat_l3 = split_breadcrumb(page_breadcrumbs)
    
    for payload in RAW_INFO_BUFFER:
        for item in payload.get("materialInfos", []):
            code = str(item.get("materialNumber", ""))
            if not code: continue
            if code not in product_db:
                product_db[code] = {
                    "name": "Unknown", "cat_l1": cat_l1, "cat_l2": cat_l2, "cat_l3": cat_l3, "brand": "Unknown",
                    "measure_1": "", "measure_2": "",
                    "price_1_unit": "Case", "price_1": "",
                    "price_2_unit": "", "price_2": "",
                    "stock": "In Stock", "last_ordered": "N/A",
                    "image_file": "No Image", "img_url": "", "changes": ""
                }
            
            desc = item.get("description") or {}
            name = desc.get("en") or desc.get("fr") or product_db[code].get("name")
            brand = (item.get("brand") or {}).get("en", "Unknown Brand")
            
            # Dynamic re-routing if the page breadcrumbs failed but the item knows where it lives
            if cat_l1 == "" and product_db[code]["cat_l1"] == "":
                hier = item.get("hierarchy", [])
                if hier and isinstance(hier, list):
                    for level in reversed(hier):
                        cat_name = level.get("name", {}).get("en")
                        if cat_name:
                            product_db[code]["cat_l1"] = cat_name
                            break
                            
            weight = item.get("baseUomWeight") or {}
            net_weight = weight.get("net", "")
            try:
                nw_float = float(net_weight)
                net_weight = f"{int(nw_float)}" if nw_float.is_integer() else f"{nw_float}"
            except Exception: pass
            
            uom = weight.get("uom", "").upper()
            if uom in ["KG", "KGM"]: uom = "Kilograms"
            elif uom in ["LB", "LBR"]: uom = "Pounds"
            elif uom in ["G", "GRM"]: uom = "Grams"
            elif uom in ["ML", "MLT"]: uom = "Milliliters"
            elif uom in ["L", "LTR"]: uom = "Liters"
            else: uom = uom.capitalize()
            
            base_uom = item.get("baseUom", "Case").capitalize()
            if base_uom in ["Cs", "Bg", "Ea"]: base_uom = {"Cs":"Case", "Bg":"Bag", "Ea":"Each"}.get(base_uom, base_uom)
            
            measure_1 = f"{net_weight} {uom}".strip()
            measure_2 = base_uom
            
            img_url = ""
            if "image" in item and item["image"]:
                for lang in ["en", "fr", "es"]:
                    img_data = item["image"].get(lang) or {}
                    if img_data.get("url"):
                        img_url = img_data["url"]
                        break
                        
            product_db[code].update({
                "name": name, 
                "brand": brand,
                "measure_1": measure_1,
                "measure_2": measure_2
            })
            
            if img_url and not product_db[code].get("img_url"):
                product_db[code]["img_url"] = img_url

    for payload in RAW_PRICE_BUFFER:
        for item in payload.get("materialPrices", []):
            code = str(item.get("materialNumber", ""))
            if not code: continue
            if code not in product_db:
                product_db[code] = {
                    "name": "Unknown", "cat_l1": cat_l1, "cat_l2": cat_l2, "cat_l3": cat_l3, "brand": "Unknown",
                    "measure_1": "", "measure_2": "",
                    "price_1_unit": "Case", "price_1": "",
                    "price_2_unit": "", "price_2": "",
                    "stock": "In Stock", "last_ordered": "N/A",
                    "image_file": "No Image", "img_url": "", "changes": ""
                }
                
            units = item.get("unitPrices") or []
            p1_u, p1_v, p2_u, p2_v = "", "", "", ""
            
            for idx, u in enumerate(units):
                if u.get("price") is not None:
                    uom_str = u.get("salesUom") or u.get("uom") or "EA"
                    uom_str = {"CS":"Case", "BG":"Bag", "EA":"Each"}.get(uom_str, uom_str.capitalize())
                    p_str = f"${float(u['price']):.2f}"
                    
                    if idx == 0:
                        p1_u, p1_v = uom_str, p_str
                    elif idx == 1:
                        p2_u, p2_v = uom_str, p_str
                        
            if p1_v:
                product_db[code].update({
                    "price_1_unit": p1_u, "price_1": p1_v,
                    "price_2_unit": p2_u, "price_2": p2_v
                })

async def dump_json_buffers(suffix=""):
    global RAW_INFO_BUFFER, RAW_PRICE_BUFFER
    try:
        if RAW_INFO_BUFFER:
            filename = f"raw_info_dump_{suffix}.json" if suffix else "raw_info_dump.json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(RAW_INFO_BUFFER, f, separators=(',', ':'))
        if RAW_PRICE_BUFFER:
            filename = f"raw_price_dump_{suffix}.json" if suffix else "raw_price_dump.json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(RAW_PRICE_BUFFER, f, separators=(',', ':'))
    except Exception as e:
        print(f"⚠️ FATAL ERROR dumping JSON buffers: {e}")

async def run_harvest(page, kb=None, state="search_results"):
    global target_total, product_db, RAW_INFO_BUFFER, RAW_PRICE_BUFFER
    global_product_db = {}
    
    product_db.clear()
    RAW_INFO_BUFFER.clear()
    RAW_PRICE_BUFFER.clear()
    
    target_total = 9999 
    harvest_success = True
    update_status("running", "Initializing Harvester...", 5, 100)
    
    for _ in range(15):
        if target_total != 9999: break
        await asyncio.sleep(0.5)
        
    all_res_target = 9999
    order_guide_target = 0
    print("🔍 Scanning DOM for product totals and breadcrumbs...")
    
    page_breadcrumbs = "Unknown Category"
    try:
        page_breadcrumbs = await page.evaluate("""() => {
            let nodes = Array.from(document.querySelectorAll('.crumb, .breadcrumb-item, gfs-breadcrumbs a, gfs-breadcrumbs span'));
            let pathParts = nodes.map(el => el.innerText.trim())
                .filter(t => t && t.toLowerCase() !== 'home' && t !== '>' && t !== '/' && !t.includes('results for'))
                .map(t => t.replace(/[>|/\\n]/g, '').trim());

            let cleanPath = [];
            pathParts.forEach(part => {
                if (cleanPath.length === 0 || cleanPath[cleanPath.length - 1] !== part) {
                    cleanPath.push(part);
                }
            });

            if (cleanPath.length > 0) return cleanPath.join(' > ');
            
            let h1 = document.querySelector('h1');
            if (h1) {
                let h1Text = h1.innerText.trim().replace(/^\\d+\\s+results for\\s+"/i, '').replace(/"$/g, '');
                return "Categories > " + h1Text;
            }
            return "Unknown Category";
        }""")
        page_breadcrumbs = page_breadcrumbs.replace('\n', '').replace('\r', '').strip()
        print(f"🏷️ Page Breadcrumbs Found: {page_breadcrumbs}")
        
        text = await page.evaluate("document.body.innerText")
        text_lower = text.lower()
        
        if "0 results" in text_lower or "no results found" in text_lower:
            print("⚠️ No products found on this page. Assuming empty category.")
            return True 
            
        all_match = re.search(r'(\d+)\s+results', text_lower)
        og_match = re.search(r'order guide only\s*\((\d+)\)', text_lower)
        
        if all_match: 
            all_res_target = int(all_match.group(1))
            print(f"📊 DOM found All Results: {all_res_target}")
        if og_match: 
            order_guide_target = int(og_match.group(1))
            print(f"📊 DOM found Order Guide: {order_guide_target}")
    except Exception as e:
        print(f"⚠️ DOM parsing error: {e}")
        
    if target_total != 9999:
        print(f"📡 Wiretap intercepted target: {target_total}")
        all_res_target = target_total
        
    if all_res_target == 9999 or order_guide_target == 0:
        print("🤖 Missing some totals. Tagging in Gemini Supervisor...")
        totals = await gemini_double_check_totals(page)
        if totals:
            if all_res_target == 9999: 
                all_res_target = int(totals.get("all_results_total", 9999))
            if order_guide_target == 0: 
                order_guide_target = int(totals.get("order_guide_total", 0))
                
    target_total = all_res_target
    print(f"🎯 Target Phase 1 Confirmed: {target_total} | Target Phase 2 Confirmed: {order_guide_target}")
    
    async def perform_scroll_loop(phase_name="Phase 1"):
        stagnant_strikes = 0
        max_strikes = 15 
        seen_in_dom = set() 
        
        cat_l1, cat_l2, cat_l3 = split_breadcrumb(page_breadcrumbs)
        
        while len(seen_in_dom) < target_total and stagnant_strikes < max_strikes:
            last_count = len(seen_in_dom)
            remaining_items = max(0, target_total - last_count)
            update_status("running", f"[{phase_name}] Scrolling... {last_count}/{target_total} (Need {remaining_items} more)", last_count, target_total)
            
            await dismiss_all_popups(page)
            await force_scroll_down(page, scrolls=1)
            merge_buffers_to_db(page_breadcrumbs)
            
            try:
                dom_data = await page.evaluate("""() => {
                    let items = {};
                    let regex = /#(\\d{5,9})/;
                    let container = document.querySelector('cdk-virtual-scroll-viewport, .scroll-container, main, .product-grid, .item-list') || document.body;
                    let elements = container.querySelectorAll('*');
                    
                    for (let el of elements) {
                        if (el.children.length > 0) continue;
                        let tagName = el.tagName.toLowerCase();
                        if (tagName === 'script' || tagName === 'style' || tagName === 'noscript') continue;
                        if (el.closest('header') || el.closest('naoo-header') || el.closest('.header') || el.closest('nav') || el.closest('.docket-header')) continue;
                        
                        let text = el.innerText || el.textContent || '';
                        if (text.includes('{') || text.includes('}')) continue;
                        
                        let match = text.match(regex);
                        if (match) {
                            let code = match[1];
                            let parent = el.parentElement;
                            let maxDepth = 12; 
                            let foundCard = false;
                            let cardNode = null;
                            
                            while (parent && maxDepth > 0) {
                                let pText = parent.innerText || '';
                                if (pText.includes(code) && (pText.includes('$') || pText.includes('Case') || pText.includes('Pack') || pText.includes('Compare') || pText.includes('Price') || pText.includes('Ordered'))) {
                                    cardNode = parent;
                                    foundCard = true;
                                    if (parent.querySelector('img')) {
                                        break;
                                    }
                                }
                                parent = parent.parentElement;
                                maxDepth--;
                            }
                            
                            let lines = [];
                            let imgUrl = "";
                            
                            if (foundCard && cardNode) {
                                lines = (cardNode.innerText || '').split('\\n').map(s => s.trim()).filter(Boolean);
                                let imgs = cardNode.querySelectorAll('img');
                                for (let img of imgs) {
                                    let src = img.src || img.getAttribute('data-src') || '';
                                    if (src && !src.startsWith('data:image') && !src.includes('svg')) {
                                        imgUrl = src;
                                        break;
                                    }
                                }
                            } else if (parent && (parent.innerText || '').split('\\n').length >= 3) {
                                lines = (parent.innerText || '').split('\\n').map(s => s.trim()).filter(Boolean);
                            }
                            
                            items[code] = { lines: lines, img: imgUrl };
                        }
                    }
                    return items;
                }""")
                
                for code, data_dict in dom_data.items():
                    if code == "722349852": continue
                    seen_in_dom.add(code) 
                    lines = data_dict.get("lines", [])
                    scraped_img = data_dict.get("img", "")
                    
                    if code not in product_db:
                        if code in global_product_db:
                            product_db[code] = dict(global_product_db[code])
                        else:
                            product_db[code] = {
                                "name": "Unknown", "cat_l1": cat_l1, "cat_l2": cat_l2, "cat_l3": cat_l3, "brand": "Unknown",
                                "measure_1": "", "measure_2": "",
                                "price_1_unit": "Case", "price_1": "",
                                "price_2_unit": "", "price_2": "",
                                "stock": "In Stock", "last_ordered": "N/A",
                                "image_file": "No Image", "img_url": "", "changes": ""
                            }
                            
                    if scraped_img and (not product_db[code].get("img_url") or "no_image" in product_db[code].get("img_url", "").lower()):
                        product_db[code]["img_url"] = scraped_img
                        
                    if len(lines) > 0:
                        ignore_labels = ["compare", "view similar items", "long term out of stock", "new", "in stock", "contact sales", "add to cart", "out of stock", "special order", "hide unavailable items", "new item coming soon"]
                        clean_lines = [l for l in lines if l.lower() not in ignore_labels and "delivery" not in l.lower()]
                        
                        ordered_lines = [l for l in clean_lines if "ordered:" in l.lower() or "ordered " in l.lower()]
                        for l in ordered_lines:
                            match = re.search(r'ordered:?\s*(.*)', l, re.IGNORECASE)
                            if match:
                                product_db[code]["last_ordered"] = match.group(1).strip()
                            clean_lines.remove(l) 
                            
                        code_idx = -1
                        for i, l in enumerate(clean_lines):
                            if f"#{code}" in l:
                                code_idx = i
                                break
                                
                        if code_idx != -1:
                            name_found = False
                            if code_idx > 0:
                                badges = ["local", "new", "compare", "view similar items", "can", "us", "add to list", "order guide"]
                                for i in range(code_idx - 1, -1, -1):
                                    if clean_lines[i].lower() not in badges and len(clean_lines[i]) > 3:
                                        if product_db[code]["name"] in ["Unknown", ""]:
                                            product_db[code]["name"] = clean_lines[i]
                                        name_found = True
                                        break
                                
                            if not name_found and product_db[code]["name"] in ["Unknown", ""] and clean_lines:
                                product_db[code]["name"] = clean_lines[0]
                                
                            code_line = clean_lines[code_idx]
                            parts = [p.strip() for p in code_line.split("|")]
                            
                            if len(parts) >= 3:
                                if product_db[code]["brand"] in ["Unknown", ""]:
                                    product_db[code]["brand"] = parts[1]
                                
                                measure_str = parts[2]
                                if "," in measure_str:
                                    m_parts = [m.strip() for m in measure_str.split(",", 1)]
                                    product_db[code]["measure_1"] = m_parts[0]
                                    product_db[code]["measure_2"] = m_parts[1]
                                else:
                                    product_db[code]["measure_1"] = measure_str
                                    
                            elif len(parts) == 2:
                                if any(char.isdigit() for char in parts[1]):
                                    measure_str = parts[1]
                                    if "," in measure_str:
                                        m_parts = [m.strip() for m in measure_str.split(",", 1)]
                                        product_db[code]["measure_1"] = m_parts[0]
                                        product_db[code]["measure_2"] = m_parts[1]
                                    else:
                                        product_db[code]["measure_1"] = measure_str
                                else:
                                    if product_db[code]["brand"] in ["Unknown", ""]:
                                        product_db[code]["brand"] = parts[1]
                        else:
                            if product_db[code]["name"] in ["Unknown", ""] and clean_lines:
                                product_db[code]["name"] = clean_lines[0]
                                
                        price_list = []
                        for i, l in enumerate(clean_lines):
                            if "$" in l and any(c.isdigit() for c in l):
                                if ":" in l: 
                                    p_parts = l.split(":", 1)
                                    unit = p_parts[0].strip()
                                    val = p_parts[1].strip()
                                else: 
                                    unit = clean_lines[i-1] if i > 0 and len(clean_lines[i-1]) <= 6 else "Case"
                                    val = l
                                price_list.append((unit, val))

                        if price_list:
                            if product_db[code]["price_1"] == "":
                                product_db[code]["price_1_unit"] = price_list[0][0]
                                product_db[code]["price_1"] = price_list[0][1]
                            if len(price_list) > 1 and product_db[code]["price_2"] == "":
                                product_db[code]["price_2_unit"] = price_list[1][0]
                                product_db[code]["price_2"] = price_list[1][1]

            except Exception as e:
                print(f"⚠️ DOM extraction error: {e}")
                
            if len(seen_in_dom) == last_count:
                stagnant_strikes += 1
                print(f"⚠️ Memory Stagnant on {phase_name}. Strike {stagnant_strikes}/{max_strikes}")
                try:
                    await page.evaluate("""() => {
                        const viewport = document.querySelector('cdk-virtual-scroll-viewport, .scroll-container');
                        if (viewport) {
                            viewport.scrollTop -= 200; 
                            setTimeout(() => { viewport.scrollTop += 1200; }, 150); 
                        } else {
                            window.scrollBy(0, -200);
                            setTimeout(() => { window.scrollBy(0, 1200); }, 150);
                        }
                    }""")
                    viewport = page.viewport_size
                    if viewport:
                        safe_x = int(viewport['width'] - 10)
                        safe_y = int(viewport['height'] * 0.5)
                        await page.mouse.click(safe_x, safe_y)
                    await page.keyboard.press("PageDown")
                    await asyncio.sleep(2.0)
                except: pass
            else: 
                stagnant_strikes = 0
                print(f"📦 [{phase_name}] Progress: {len(seen_in_dom)} / {target_total} products mapped. ({max(0, target_total - len(seen_in_dom))} left to go!)")
                
    try:
        print("🚀 Starting Phase 1: Scraping 'All Results' tab...")
        await perform_scroll_loop(phase_name="Phase 1")
        await dump_json_buffers("phase1")
        merge_buffers_to_db(page_breadcrumbs)
        global_product_db.update(product_db)
        
        if order_guide_target > 0 and kb is not None:
            print(f"\n🔄 Initiating Phase 2: Switching to 'Order Guide Only' tab to fetch {order_guide_target} items...")
            await page.evaluate("""() => {
                window.scrollTo({ top: 0, behavior: 'smooth' });
                document.documentElement.scrollTop = 0;
                document.body.scrollTop = 0;
                const elements = document.querySelectorAll('cdk-virtual-scroll-viewport, .scroll-container, main, .product-grid, .item-list');
                for (let el of elements) {
                    el.scrollTop = 0;
                    el.dispatchEvent(new Event('scroll'));
                }
            }""")
            
            await asyncio.sleep(2.0)
            for _ in range(5):
                await page.keyboard.press("PageUp")
                await asyncio.sleep(0.2)
            for _ in range(3):
                await page.keyboard.press("Home")
                await asyncio.sleep(0.5)
                
            tab_clicked = False
            try:
                selectors_to_try = [
                    "text=/Order Guide Only/i",
                    "[role='tab']:has-text('Order Guide')",
                    "div.tab-header:has-text('Order Guide')",
                    "//div[contains(translate(text(), 'ORDER', 'order'), 'order guide')]"
                ]
                for sel in selectors_to_try:
                    try:
                        tab_loc = page.locator(sel).first
                        await tab_loc.wait_for(state="attached", timeout=2000)
                        await tab_loc.scroll_into_view_if_needed()
                        await asyncio.sleep(0.5)
                        await tab_loc.click(force=True)
                        tab_clicked = True
                        print(f"✅ Switched tabs via direct DOM Locator successfully using: {sel}")
                        await asyncio.sleep(2)
                        await dismiss_all_popups(page) 
                        break
                    except Exception:
                        continue
            except Exception as e:
                print(f"⚠️ Direct DOM clicker failed: {e}")
                pass
                
            if not tab_clicked:
                tab_clicked = await teach_and_click(page, kb, state, "order_guide_tab", action="click")
                
            if tab_clicked:
                print("✅ Wiping memory buffers for Phase 2 isolation...")
                product_db.clear()
                RAW_INFO_BUFFER.clear()
                RAW_PRICE_BUFFER.clear()
                target_total = order_guide_target
                await asyncio.sleep(5) 
                merge_buffers_to_db(page_breadcrumbs) 
                await perform_scroll_loop(phase_name="Phase 2")
                await dump_json_buffers("phase2")
                merge_buffers_to_db(page_breadcrumbs)
                
                for code, data in product_db.items():
                    if "[ORDER GUIDE]" not in data["name"]:
                        data["name"] = f"⭐ [ORDER GUIDE] {data['name']}"
                    global_product_db[code] = data
                    
        total_collected = len(global_product_db)
        expected_minimum = int(all_res_target * 0.90) 
        if total_collected < expected_minimum and all_res_target != 9999:
            print(f"\n⚠️ [HARVEST INCOMPLETE] Only collected {total_collected} out of expected ~{all_res_target} items.")
            harvest_success = False
        else:
            print(f"\n✅ [HARVEST COMPLETE] Successfully collected {total_collected} items.")
            harvest_success = True
            
    except Exception as e:
        print(f"⚠️ HARVEST WARNING: Sequence interrupted ({e}). Saving retrieved data...")
        harvest_success = False
        
    finally:
        update_status("processing", "Saving to SQLite DB & Generating Master Excel...", 90, 100)
        
        keys_to_delete = []
        for k, v in global_product_db.items():
            if v["name"] in ["Unknown", ""] and v["price_1"] == "":
                keys_to_delete.append(k)
        for k in keys_to_delete:
            del global_product_db[k]
        if keys_to_delete:
            print(f"👻 Cleaned up {len(keys_to_delete)} Virtual Scroller ghost artifacts.")
            
        product_db.clear()
        product_db.update(global_product_db)
        await download_pending_images(page)
        process_changes()
        generate_excel_master()
        generate_html_dashboard()
        
    return harvest_success